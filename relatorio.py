"""
relatorio.py — Gera Excel com seriais pendentes
Recebe a lista diretamente da memória (sem banco de dados).
"""

import os
from datetime import datetime


def gerar_excel(registros: list[dict], destino: str) -> str:
    """
    Gera o arquivo Excel com os seriais pendentes.

    registros: lista de dicts gerada pelo app:
        [{"Serial": "...", "Nome": "...", "Loja": "...", "Status": "..."}, ...]

    destino: caminho completo do arquivo .xlsx a ser salvo.

    Retorna o caminho do arquivo gerado.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError(
            "Biblioteca 'openpyxl' não instalada.\n"
            "Execute: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pendentes"

    # ── Estilos ──────────────────────────────────────────────────────────────
    # Cores alteradas para tons de Laranja/Âmbar (tema de pendência)
    fill_titulo    = PatternFill("solid", fgColor="B85C00")
    fill_cabecalho = PatternFill("solid", fgColor="E67300")
    fill_par       = PatternFill("solid", fgColor="FFF5E6")
    fill_impar     = PatternFill("solid", fgColor="FFFFFF")
    fill_rodape    = PatternFill("solid", fgColor="FFF9F2")

    fonte_titulo    = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    fonte_cabecalho = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    fonte_normal    = Font(name="Calibri", size=10)
    fonte_serial    = Font(name="Consolas", size=10)
    fonte_rodape    = Font(name="Calibri", size=9, italic=True, color="888888")
    fonte_total     = Font(name="Calibri", size=10, bold=True, color="B85C00")

    borda = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    centro = Alignment(horizontal="center", vertical="center")
    esq    = Alignment(horizontal="left",   vertical="center")

    # ── Linha 1: Título ───────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 32
    cel = ws["A1"]
    cel.value     = "⏳  Seriais Pendentes — Relatório de Auditoria"
    cel.font      = fonte_titulo
    cel.fill      = fill_titulo
    cel.alignment = centro

    # ── Linha 2: Metadados ────────────────────────────────────────────────────
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 18
    cel = ws["A2"]
    cel.value = (
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   "
        f"Total de seriais pendentes: {len(registros)}"
    )
    cel.font      = fonte_rodape
    cel.alignment = centro
    cel.fill      = fill_rodape

    # ── Linha 3: Cabeçalho ────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 22
    colunas = ["#", "Serial", "Nome", "Loja", "Status"]
    for col, titulo in enumerate(colunas, start=1):
        cel           = ws.cell(row=3, column=col, value=titulo)
        cel.font      = fonte_cabecalho
        cel.fill      = fill_cabecalho
        cel.alignment = centro
        cel.border    = borda

    # ── Dados ─────────────────────────────────────────────────────────────────
    for i, reg in enumerate(registros, start=1):
        row  = i + 3
        fill = fill_par if i % 2 == 0 else fill_impar
        ws.row_dimensions[row].height = 18

        # Nº
        cel = ws.cell(row=row, column=1, value=i)
        cel.font = fonte_normal; cel.fill = fill
        cel.alignment = centro;  cel.border = borda

        # Serial
        cel = ws.cell(row=row, column=2, value=reg.get("Serial", ""))
        cel.font = fonte_serial; cel.fill = fill
        cel.alignment = esq;     cel.border = borda

        # Nome
        cel = ws.cell(row=row, column=3, value=reg.get("Nome", ""))
        cel.font = fonte_normal; cel.fill = fill
        cel.alignment = esq;     cel.border = borda

        # Loja
        cel = ws.cell(row=row, column=4, value=reg.get("Loja", ""))
        cel.font = fonte_normal; cel.fill = fill
        cel.alignment = centro;  cel.border = borda

        # Status
        cel = ws.cell(row=row, column=5, value=reg.get("Status", "Pendente"))
        cel.font = fonte_normal; cel.fill = fill
        cel.alignment = centro;  cel.border = borda

    # ── Linha de total ────────────────────────────────────────────────────────
    row_total = len(registros) + 4
    ws.row_dimensions[row_total].height = 20
    ws.merge_cells(f"A{row_total}:D{row_total}")

    cel = ws.cell(row=row_total, column=1,
                  value="Total de seriais pendentes:")
    cel.font = fonte_total; cel.fill = fill_rodape
    cel.alignment = esq;    cel.border = borda

    cel = ws.cell(row=row_total, column=5, value=len(registros))
    cel.font = fonte_total; cel.fill = fill_rodape
    cel.alignment = centro; cel.border = borda

    # ── Largura das colunas ───────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 15

    ws.freeze_panes = "A4"

    wb.save(destino)
    return destino